import csv
import pandas as pd
from django.shortcuts import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from polls.filters import NewsFilter
from rest_framework import generics, filters
from openpyxl import Workbook
from rest_framework.views import APIView
from polls.serializers import NewsSerializer,FilterDataSerializer
from polls.models import News, FilterData
from rest_framework.response import Response
from rest_framework import status




class FilterDataView(APIView):
    def get(self, request):
        movie1 = FilterData.objects.all()
        serializer = FilterDataSerializer(movie1)
        return Response(serializer.data)

    def post(self,request):
        site = request.data.pop('sitename')
        word = request.data.pop('titlecontent')
        data = pd.read_csv('./polls/news.csv')
        data = data.loc[data.isin([site, word])]
        data.to_csv('./polls/importnews.csv')
        return Response(status=status.HTTP_200_OK)


class NewsList(generics.ListCreateAPIView):
    queryset = News.objects.all()
    serializer_class = NewsSerializer

    filter_backends = (DjangoFilterBackend,
                       filters.SearchFilter,
                       filters.OrderingFilter)

    filter_class = NewsFilter
    search_fields = ('pagefrom', 'title')


# def dowLoadCSV(request):
#     items = Movies.objects.all()
#     response = HttpResponse()
#     response['Content-Disposition'] = 'attachment; filename="news.csv"'
#
#     writer  = csv.writer(response, delimiter=',')
#     writer.writerow(['id', 'title', 'director', 'producer', 'rating'])
#
#     for obj in items:
#         writer.writerow([obj.id, obj.title, obj.director, obj.producer, obj.rating])
#
#     return response
#
#
# def dowLoadXLSX(request):
#     items = Movies.objects.all()
#
#     response = HttpResponse()
#     response['Content-Disposition'] = 'attachment; filename="news.xlsx"'
#
#     workbook = Workbook()
#     worksheet = workbook.active
#     worksheet.title = 'Movies'
#
#     columns = [
#         'ID',
#         'Title',
#         'Director',
#         'Producer',
#         'Rating',
#     ]
#     row_num = 1
#
#     for col_num, column_title in enumerate(columns, 1):
#         cell = worksheet.cell(row=row_num, column=col_num)
#         cell.value = column_title
#
#     for movie in items:
#         row_num += 1
#         row = [
#             movie.id,
#             movie.title,
#             movie.director,
#             movie.producer,
#             movie.rating,
#         ]
#         for col_num, cell_value in enumerate(row, 1):
#             cell = worksheet.cell(row=row_num, column=col_num)
#             cell.value = cell_value
#
#     workbook.save(response)
#
#     return response
#
